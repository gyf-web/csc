import csv
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "文献整理.xlsx"
INPUT = ROOT / "supplement_candidates_verified.csv"
OUTPUT = ROOT / "文献整理_补充检索.xlsx"
REPORT = ROOT / "补充文献检索.md"


ASSESSMENTS = {
    "10.1088/1748-9326/ab9e99": ("GEDI复杂性概念与指标", "高", "核心纳入", "直接讨论利用GEDI从太空制图冠层结构多样性，是WSCI和全球复杂性研究的重要前置文献。"),
    "10.3390/rs14153615": ("GEDI结构变异验证", "高", "核心纳入", "直接评估GEDI量化澳大利亚温带森林结构变异的能力，可补足森林类型与区域验证证据。"),
    "10.1038/s41597-024-03668-4": ("GEDI多尺度栅格产品", "高", "核心纳入", "将超过70亿个GEDI观测汇总为36项植被结构指标的多分辨率栅格，为尺度效应和连续制图提供数据基础。"),
    "10.1016/j.rse.2024.114174": ("干扰与重复足迹", "高", "核心纳入", "利用重复GEDI足迹量化热带森林干扰后的结构变化，补充时间变化和退化监测方向。"),
    "10.1016/j.rse.2023.113703": ("结构多样性与碳储量", "高", "核心纳入", "联合GEDI与森林清查分析结构多样性、物种多样性和地上碳储量关系，属于生态功能应用。"),
    "10.1002/fee.2585": ("GEDI结构生态监测综述", "高", "核心纳入", "面向生态监测总结GEDI冠层结构产品与应用，可用于绪论、指标说明和应用框架。"),
    "10.1016/j.rse.2024.114446": ("GEDI结构与鸟类多样性", "高", "核心纳入", "直接检验GEDI植被结构与北美不同鸟类功能群多样性的关系，扩展生物多样性应用。"),
    "10.1016/j.gloenvcha.2022.102621": ("保护地生态结构评价", "高", "核心纳入", "利用GEDI评价保护地碳储量和生态结构，体现复杂性相关结构指标在保护成效评价中的应用。"),
    "10.3390/rs15081969": ("GEDI与Sentinel连续制图", "高", "核心纳入", "基于GEDI、Sentinel-1和Sentinel-2生成德国森林结构产品，可与现有GEDI-fusion研究比较。"),
    "10.3390/rs13245105": ("GEDI与Sentinel连续制图", "高", "核心纳入", "在巴拉圭查科地区以Sentinel-1/2外推GEDI结构指标，是区域连续制图和GEE实现的重要案例。"),
    "10.1016/j.rse.2025.114930": ("结构复杂性与碳水通量", "高", "已有记录更新", "原表已收录该文但年份、期刊和DOI未说明；现已核验正式题录。"),
    "10.1088/3049-4753/ae2c01": ("GEDI与SAR全球复杂性制图", "高", "核心纳入", "以GEDI约束并融合SAR进行全球森林结构复杂性连续制图，直接对应多源融合与全球外推前沿。"),
    "10.21105/joss.08593": ("GEDI开源数据工具", "中高", "方法纳入", "提供GEDI L2A/L2B及L4A/L4C处理和供给工具，适合写入公开代码与可复现性部分。"),
    "10.1016/j.rse.2022.113280": ("冠层熵方法", "中高", "方法纳入", "提出跨平台LiDAR冠层熵方法，是GEDI冠层熵及结构复杂性指标构建的重要方法基础。"),
    "10.1038/s41467-020-20767-z": ("全球复杂性格局", "中高", "理论纳入", "基于地基激光雷达揭示全球结构复杂性格局和气候控制，为GEDI全球结果提供理论参照。"),
    "10.1073/pnas.2506750122": ("复杂性生态机制", "中高", "理论纳入", "阐释生物多样性通过冠层结构复杂性提高生产力的机制，可支撑GEDI复杂性生态意义讨论。"),
    "10.1111/1365-2745.70241": ("次生演替与恢复", "中高", "应用纳入", "研究热带次生演替中结构复杂性的恢复轨迹，可为GEDI恢复监测提出可检验假设。"),
    "10.1016/j.ecolind.2025.113085": ("多平台指标比较", "中高", "方法纳入", "比较多平台、多传感器结构复杂性指标，有助于讨论GEDI指标的跨平台一致性和验证。"),
    "10.1016/j.tfp.2025.100954": ("复杂性时间变化", "中", "拓展阅读", "使用激光扫描捕捉结构复杂性发展趋势，可借鉴时间变化指标和监测设计。"),
    "10.1080/22797254.2024.2417905": ("双时相复杂性变化", "中", "拓展阅读", "以双时相机载激光雷达量化复杂性变化，可作为GEDI重复观测研究的方法参照。"),
}


def normalize_doi(value):
    return (value or "").strip().lower().replace("https://doi.org/", "")


def short_abstract(text, limit=360):
    text = " ".join((text or "").split())
    if not text:
        return "未获取摘要；相关性判断基于题名和核验元数据。"
    return text if len(text) <= limit else text[:limit].rstrip() + "..."


with INPUT.open(encoding="utf-8-sig", newline="") as fh:
    candidates = list(csv.DictReader(fh))

wb = load_workbook(SOURCE)
main = wb["文献编码表"]
headers = {cell.value: cell.column for cell in main[1]}
existing_dois = {
    normalize_doi(main.cell(row, headers["DOI"]).value)
    for row in range(2, main.max_row + 1)
    if normalize_doi(main.cell(row, headers["DOI"]).value) not in {"", "未说明"}
}

corrections = []
for row in range(2, main.max_row + 1):
    title = str(main.cell(row, headers["标题"]).value or "")
    if title.startswith("Upscaling eddy covariance measurements"):
        main.cell(row, headers["年份"]).value = "2025"
        main.cell(row, headers["期刊"]).value = "Remote Sensing of Environment"
        main.cell(row, headers["DOI"]).value = "10.1016/j.rse.2025.114930"
        corrections.append("补全 Bu & Xiao (2025) 的年份、期刊和 DOI。")
    elif title == "Mapping global forest canopy height through integration of GEDI and Landsat data":
        main.cell(row, headers["DOI"]).value = "10.1016/j.rse.2020.112165"
        corrections.append("补全 Potapov et al. 全球冠层高度论文 DOI。")

if "补充检索文献" in wb.sheetnames:
    del wb["补充检索文献"]
ws = wb.create_sheet("补充检索文献")
columns = [
    "序号", "标题", "作者", "年份", "期刊", "DOI", "主题类别", "相关性",
    "纳入建议", "与研究主题的关系", "是否已在原表", "摘要摘录/检索依据",
    "开放获取", "OpenAlex被引次数", "题录链接", "元数据来源", "检索日期", "核验状态",
]
ws.append(columns)

for idx, item in enumerate(candidates, 1):
    doi = normalize_doi(item["DOI"])
    category, relevance, recommendation, relation = ASSESSMENTS[doi]
    already = "是" if doi in existing_dois or recommendation == "已有记录更新" else "否"
    ws.append([
        idx,
        item["Title"].replace("<scp>", "").replace("</scp>", ""),
        item["Authors"],
        str(item["Year"]),
        item["Journal"],
        doi,
        category,
        relevance,
        recommendation,
        relation,
        already,
        short_abstract(item.get("Abstract")),
        "是" if item["OA"].lower() == "true" else "否",
        int(item["Citations"] or 0),
        item["Landing"],
        "OpenAlex（DOI级元数据核验）",
        "2026-06-22",
        "DOI、题名、作者、年份和期刊已核验；研究细节待全文复核",
    ])

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(name="Arial", bold=True, color="FFFFFF")
body_font = Font(name="Arial", size=10)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = [8, 58, 48, 10, 30, 31, 24, 11, 16, 58, 15, 62, 12, 16, 42, 30, 14, 42]
for idx, width in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

if "补充检索说明" in wb.sheetnames:
    del wb["补充检索说明"]
note = wb.create_sheet("补充检索说明")
notes = [
    ("项目", "内容"),
    ("研究主题", "GEDI数据在森林结构复杂性反演中的应用"),
    ("检索日期", "2026-06-22"),
    ("检索工作流", "nature-academic-search / multi-source-search"),
    ("T1来源", "OpenAlex聚合的Crossref及出版社DOI元数据；Crossref接口遇到429限流"),
    ("T2来源", "Semantic Scholar尝试补充摘要，但接口返回429；未采用未经核验结果"),
    ("核心检索词", "GEDI; Global Ecosystem Dynamics Investigation; forest structural complexity; canopy structural complexity; vegetation structure; structural diversity"),
    ("纳入标准", "直接使用GEDI研究森林三维/垂直结构，或为复杂性指标、尺度转换、连续制图、生态功能与开源工具提供强支撑"),
    ("排重规则", "优先按标准化DOI排重；DOI缺失时应按题名和第一作者复核"),
    ("重要限制", "新增文献未提供本地PDF，研究区、筛选阈值、模型与评价指标尚未逐篇全文编码；不得把本表摘要判断视为全文证据"),
    ("被引次数", "OpenAlex动态值，仅用于辅助排序，检索日后可能变化"),
]
for row in notes:
    note.append(row)
for cell in note[1]:
    cell.fill = header_fill
    cell.font = header_font
note.column_dimensions["A"].width = 22
note.column_dimensions["B"].width = 110
for row in note.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUTPUT)

core = [c for c in candidates if ASSESSMENTS[normalize_doi(c["DOI"])][2] == "核心纳入"]
methods = [c for c in candidates if ASSESSMENTS[normalize_doi(c["DOI"])][2] in {"方法纳入", "理论纳入", "应用纳入"}]
report = f"""# GEDI森林结构复杂性补充文献检索

## 检索结果

本轮在原有27篇文献基础上核验20篇候选文献，其中建议新增核心文献{len(core)}篇，方法、理论或应用支撑文献{len(methods)}篇，拓展阅读2篇，已有记录题录更新1篇。所有候选文献均完成DOI级题名、作者、年份和期刊核验；未获得全文的字段没有进行推断性编码。

## 最优先补充

1. Schneider et al. (2020)：GEDI冠层结构多样性空间制图的早期关键论文，DOI: 10.1088/1748-9326/ab9e99。
2. Burns et al. (2024)：基于超过70亿个GEDI观测生成36项多分辨率植被结构指标，DOI: 10.1038/s41597-024-03668-4。
3. de Conto et al. (2025)：GEDI与SAR深度融合的全球森林结构复杂性连续制图，DOI: 10.1088/3049-4753/ae2c01。
4. Holcomb et al. (2024)：利用重复GEDI足迹监测热带森林干扰效应，DOI: 10.1016/j.rse.2024.114174。
5. Crockett et al. (2023)：GEDI结构多样性、物种多样性与森林碳储量关系，DOI: 10.1016/j.rse.2023.113703。
6. Kacic et al. (2021, 2023)：GEDI与Sentinel连续结构制图案例，DOI: 10.3390/rs13245105；10.3390/rs15081969。

## 对原表的题录修正

- Bu & Xiao论文已核验为2025年《Remote Sensing of Environment》，DOI: 10.1016/j.rse.2025.114930。
- “Mapping global forest canopy height through integration of GEDI and Landsat data”的DOI补全为10.1016/j.rse.2020.112165。

## 使用提醒

“补充检索文献”工作表适合作为下一批PDF下载与全文编码清单。涉及研究区、GEDI质量筛选、复杂性计算、模型和精度指标的字段，应在取得全文或补充材料后再并入主编码表。
"""
REPORT.write_text(report, encoding="utf-8")

print(OUTPUT)
print(REPORT)
print("Corrections:", len(corrections))
