from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "文献整理.xlsx"
OUT_MD = ROOT / "literature_summary.md"


rows = [
    {
        "PDF文件": "A dataset on the structural diversity of European forests.pdf",
        "标题": "A dataset on the structural diversity of European forests",
        "作者": "Marco Girardello; Gonzalo Oton; Matteo Piccardo; Mark Pickering; Agata Elia; Guido Ceccherini; Mariano Garcia; Mirco Migliavacca; Alessandro Cescatti",
        "年份": "2026",
        "期刊": "Earth System Science Data",
        "DOI": "10.5194/essd-18-2667-2026",
        "研究区": "欧洲森林，覆盖温带、地中海和大陆性区域",
        "研究对象": "森林结构多样性/冠层结构复杂性连续制图",
        "数据源": "GEDI L2A相对高度、GEDI L2B冠层覆盖度；多光谱与SAR遥感特征",
        "GEDI高精度足迹点筛选方法": "PDF明示GEDI为稀疏足迹并需足够观测支撑不同空间单元；具体quality_flag/degrade_flag阈值未说明",
        "森林复杂性的计算方法": "基于GEDI RH和冠层覆盖度计算8个结构多样性指标，包括垂直变异、偏度、峰度、Shannon-Weaver指数、Rao二次熵、凸包体积等",
        "空间分辨率": "1 km、5 km、10 km",
        "时间范围": "论文数据集面向欧洲森林；具体GEDI观测年份在抽取文本中未明确说明",
        "方法模型": "Random Forest；后向逐步特征选择；遥感纹理/光谱/SAR特征外推GEDI结构指标",
        "评价指标": "MSE；模型稳健性描述",
        "主要结论": "该文生成欧洲森林8类结构多样性指标数据集，认为GEDI与多源遥感结合可填补稀疏足迹到连续制图的缺口。",
        "局限性": "GEDI足迹稀疏；空间分辨率与样本稳定性之间存在权衡，细分辨率不确定性更高。",
        "相关性": "高",
        "相关性判断依据": "直接使用GEDI反演/制图森林结构多样性，是研究主题的核心文献。",
    },
    {
        "PDF文件": "Application Status and Prospect of Lidar in Forest Ecosystem Monitoring and Simulation.pdf",
        "标题": "激光雷达在森林生态系统监测模拟中的应用现状与展望",
        "作者": "郭庆华; 刘瑾; 陶胜利; 薛宝林; 李乐; 徐光彩; 李文楷; 吴芳芳; 李玉美; 陈琳海; 庞树鑫",
        "年份": "2014",
        "期刊": "科学通报 / Chinese Science Bulletin",
        "DOI": "10.1360/972013-592",
        "研究区": "综述文献，未限定单一研究区",
        "研究对象": "LiDAR在森林生态系统结构、功能参数和模拟中的应用",
        "数据源": "地基、机载和星载LiDAR相关研究",
        "GEDI高精度足迹点筛选方法": "不适用（论文早于GEDI正式运行）",
        "森林复杂性的计算方法": "综述点云、波形、高度/密度/叶面积等结构参数提取方法；未提出GEDI复杂性指标",
        "空间分辨率": "未说明",
        "时间范围": "截至2014年前相关研究",
        "方法模型": "综述",
        "评价指标": "未说明",
        "主要结论": "LiDAR可在样地到全球尺度提供森林三维结构信息，是生态系统监测、碳循环和生物多样性研究的重要技术支撑。",
        "局限性": "综述指出传统调查尺度外推困难，LiDAR推广仍受成本、尺度转换和多源协同限制。",
        "相关性": "中",
        "相关性判断依据": "不涉及GEDI实证，但可作为LiDAR森林结构复杂性背景文献。",
    },
    {
        "PDF文件": "Canopy structural complexity promotes ecosystem multifunctionality in a temperate forest.pdf",
        "标题": "Canopy structural complexity promotes ecosystem multifunctionality in a temperate forest",
        "作者": "Xiaoxia Yi; Xin Yu; Wangming Zhou; Yuewei Tong; Xiaoke Zhang; Xiangcheng Mi",
        "年份": "2026",
        "期刊": "Plant Diversity",
        "DOI": "10.1016/j.pld.2026.05.016",
        "研究区": "美国弗吉尼亚州Smithsonian Conservation Biology Institute 25.6 ha温带森林动态样地",
        "研究对象": "冠层结构复杂性与生态系统多功能性",
        "数据源": "机载LiDAR；重复树木清查数据",
        "GEDI高精度足迹点筛选方法": "不适用（未使用GEDI）",
        "森林复杂性的计算方法": "冠层覆盖度、平均外冠层高度（MOCH）、FHD；FHD用冠层叶面积垂直剖面的Shannon diversity计算",
        "空间分辨率": "CHM为1 m × 1 m；分析样方尺度未说明",
        "时间范围": "样地2008年建立；具体LiDAR年份未说明",
        "方法模型": "结构方程模型（SEM）；回归模型",
        "评价指标": "pseudo-R2；标准化路径系数；显著性水平",
        "主要结论": "水平和垂直冠层结构复杂性与地上生物量、生产力和整体多功能性呈正相关。",
        "局限性": "作者指出生态功能覆盖仍有限，未来需纳入更新、分解、土壤呼吸等更多过程并跨更广森林梯度检验。",
        "相关性": "中",
        "相关性判断依据": "不使用GEDI，但提供复杂性指标及生态功能机制，可用于理论章节。",
    },
    {
        "PDF文件": "Characterizing the structural complexity of.pdf",
        "标题": "Characterizing the structural complexity of the Earth's forests with spaceborne lidar",
        "作者": "Tiago de Conto; John Armston; Ralph Dubayah",
        "年份": "2024",
        "期刊": "Nature Communications",
        "DOI": "10.1038/s41467-024-52468-2",
        "研究区": "近全球森林",
        "研究对象": "全球森林三维结构复杂性；GEDI Waveform Structural Complexity Index（WSCI）",
        "数据源": "GEDI spaceborne lidar；ALS点云训练/验证数据",
        "GEDI高精度足迹点筛选方法": "algorithm_run_flag=1；degrade_flag=0；landsat_water_persistence<10且urban_proportion<50；rx_maxamp>8×sd_corrected；sensitivity>0.95（热带>0.98）；pft_class为树木相关类型",
        "森林复杂性的计算方法": "以ALS点云CE_XYZ三维复杂性为目标，使用GEDI RH metrics训练WSCI；结合SHAP/PCA解释不同垂直层贡献",
        "空间分辨率": "GEDI足迹尺度（约25 m）；部分汇总到10 km尺度用于可视化/分析",
        "时间范围": "GEDI任务期；具体年份未说明",
        "方法模型": "机器学习回归；SHAP解释；PCA",
        "评价指标": "垂直复杂性R2=0.75、RMSE=8.8%；水平复杂性R2=0.40、RMSE=9.5%",
        "主要结论": "热带森林包含多数高复杂性观测；温带森林中少于20%达到热带复杂性中位水平；WSCI提供可解释的单一结构复杂性指标。",
        "局限性": "WSCI只能通过与垂直RH剖面的关联间接推断足迹内水平复杂性。",
        "相关性": "高",
        "相关性判断依据": "直接提出GEDI反演森林结构复杂性的全球指标，是核心方法文献。",
    },
    {
        "PDF文件": "Ecosphere - 2023 - Atkins - Integrating forest structural diversity measurement into ecological research.pdf",
        "标题": "Integrating forest structural diversity measurement into ecological research",
        "作者": "Jeff W. Atkins et al.",
        "年份": "2023",
        "期刊": "Ecosphere",
        "DOI": "10.1002/ecs2.4633",
        "研究区": "方法与工具论文，未限定单一研究区",
        "研究对象": "森林结构多样性测量体系及其生态研究整合",
        "数据源": "文献与结构测量方法；涉及LiDAR/遥感/样地方法",
        "GEDI高精度足迹点筛选方法": "不适用（非GEDI实证研究）",
        "森林复杂性的计算方法": "讨论高度、覆盖、体积、异质性、多样性等结构维度及指标；具体计算依研究目标而定",
        "空间分辨率": "未说明",
        "时间范围": "未说明",
        "方法模型": "方法框架/综述",
        "评价指标": "未说明",
        "主要结论": "将结构多样性纳入生态研究有助于解释生物多样性、生态系统功能和栖息地结构。",
        "局限性": "不是GEDI反演研究，需与GEDI数据产品结合后才能用于你的主题。",
        "相关性": "中",
        "相关性判断依据": "提供结构多样性定义和测量框架，可支撑指标体系章节。",
    },
    {
        "PDF文件": "Effects of forest canopy structural complexity on productivity and its.pdf",
        "标题": "Effects of forest canopy structural complexity on productivity and its stability in the Qinling Mountains: Insights into environmental dependence",
        "作者": "Chenlu Li; Fei Yu; Tiezhu Jiao; Jiangbo Yan; Xiaohua Feng; Liang Zhao; Arshad Ali; Huiwen Li; Zuoqiang Yuan",
        "年份": "2026",
        "期刊": "Trees, Forests and People",
        "DOI": "10.1016/j.tfp.2026.101197",
        "研究区": "中国秦岭山脉南北坡",
        "研究对象": "冠层结构复杂性对NPP及其稳定性的影响",
        "数据源": "UAV LiDAR；环境因子；NPP/稳定性数据",
        "GEDI高精度足迹点筛选方法": "不适用（研究使用UAV LiDAR，未使用GEDI）",
        "森林复杂性的计算方法": "LAI、平均冠层高度（MCH）和冠层熵（CE）等LiDAR结构属性",
        "空间分辨率": "点云平均密度约94 points/m2；样方尺度未说明",
        "时间范围": "未说明",
        "方法模型": "空间采样；统计/模型分析",
        "评价指标": "解释率；显著性；文摘明示结构与环境共同解释平均生产力53%变异",
        "主要结论": "秦岭南坡冠层结构复杂性和生产力稳定性更高；结构属性与环境因子共同影响生产力。",
        "局限性": "非GEDI研究，区域尺度较小；抽取文本未见对GEDI外推的实证验证。",
        "相关性": "中",
        "相关性判断依据": "提供森林复杂性生态效应和冠层熵方法，但不能直接回答GEDI反演问题。",
    },
    {
        "PDF文件": "Enhancing ecosystem productivity and stability withincreasing canopy structural complexity inglobal forests.pdf",
        "标题": "Enhancing ecosystem productivity and stability with increasing canopy structural complexity in global forests",
        "作者": "Xiaoqiang Liu; Yuhao Feng; Tianyu Hu; Yue Luo; Xiaoxia Zhao; Jin Wu; Eduardo E. Maeda; Weiming Ju; Lingli Liu; Qinghua Guo; Yanjun Su",
        "年份": "2024",
        "期刊": "Science Advances",
        "DOI": "10.1126/sciadv.adl1947",
        "研究区": "全球森林",
        "研究对象": "全球森林冠层结构复杂性与生产力/稳定性的关系",
        "数据源": "GEDI足迹（文中称超过4亿个）；机载LiDAR样地；MODIS GPP/NDVI；GOSIF；气候、土壤、物种丰富度和森林年龄数据",
        "GEDI高精度足迹点筛选方法": "抽取文本未列出quality_flag等具体阈值；明示使用大量GEDI足迹并与机载LiDAR数据集成",
        "森林复杂性的计算方法": "采用冠层熵（canopy entropy）量化CSC，并用GEDI与4000个森林样地机载LiDAR数据建立近全球CSC估计",
        "空间分辨率": "GEDI足迹估计汇总到0.005°、0.01°、0.05°",
        "时间范围": "基准年2021；稳定性使用20年以上MODIS NDVI",
        "方法模型": "随机森林分析影响因子；全球关系分析",
        "评价指标": "GEDI估计冠层熵与机载LiDAR独立估计对比R2=0.58，normalized RMSE=12.35%",
        "主要结论": "全球范围内CSC与生态系统生产力和稳定性多呈正相关；管理林CSC较低但CSC对生产力/稳定性的增强效应更强。",
        "局限性": "不同生物群区主导因子不同；GEDI估计精度仍有限，需考虑生物、气候和土壤因素的共同作用。",
        "相关性": "高",
        "相关性判断依据": "直接用GEDI反演全球冠层结构复杂性并关联生态功能。",
    },
    {
        "PDF文件": "Evaluating GEDI data fusions for continuous characterizations of forest wildlife habitat.pdf",
        "标题": "Evaluating GEDI data fusions for continuous characterizations of forest wildlife habitat",
        "作者": "Jody C. Vogeler; Patrick A. Fekety; Lisa Elliott; Neal C. Swayze; Steven K. Filippelli; Brent Barry; Joseph D. Holbrook; Kerri T. Vierling",
        "年份": "2023",
        "期刊": "Frontiers in Remote Sensing",
        "DOI": "10.3389/frsen.2023.1196554",
        "研究区": "美国西部六州：Colorado, Wyoming, Idaho, Oregon, Washington, Montana",
        "研究对象": "森林野生动物生境相关结构指标的GEDI融合连续制图",
        "数据源": "GEDI；Landsat；Sentinel-1 SAR；扰动、地形和生物气候变量；ALS验证样本；啄木鸟样点",
        "GEDI高精度足迹点筛选方法": "夏季GEDI shots；solar elevation<0；degrade_flag=0；quality_flag=1；beam sensitivity>=0.95；仅full power beams；空间均衡抽稀，每年先抽150000点、60×60 km瓦片内保留225个最大距离点",
        "森林复杂性的计算方法": "预测8个GEDI结构指标：RH98、RH75、RH50、FHD、cover、PAVD 5-10 m、PAVD>20 m、PAVD>40 m；FHD/PAVD表征垂直复杂性",
        "空间分辨率": "30 m连续预测图",
        "时间范围": "GEDI模型年2019-2020；回推2016-2018",
        "方法模型": "Random Forest GEDI-fusion；比较Combined与Landsat/Topo/Bio等模型；野生动物生境模型",
        "评价指标": "测试R2=0.36-0.76；RMSE/Bias；啄木鸟模型AUC=0.76-0.87",
        "主要结论": "GEDI融合产品可将足迹尺度结构指标扩展到30 m连续空间，并能支持不同结构需求的野生动物生境建模。",
        "局限性": "作者强调需进一步测试不同地区/森林条件下性能；ALS与GEDI系统差异限制部分PAVD验证。",
        "相关性": "高",
        "相关性判断依据": "是GEDI结构指标连续制图与生态应用的直接案例。",
    },
    {
        "PDF文件": "Evaluating GEDI for quantifying forest structure across a gradient of degradation in Amazonian rainforest.pdf",
        "标题": "Evaluating GEDI for quantifying forest structure across a gradient of degradation in Amazonian rainforests",
        "作者": "Emily L. Doyle; Hugh A. Graham; Chris A. Boulton; Timothy M. Lenton; Ted R. Feldpausch; Andrew M. Cunliffe",
        "年份": "2025",
        "期刊": "Environmental Research Letters",
        "DOI": "10.1088/1748-9326/adc752",
        "研究区": "巴西亚马逊不同退化梯度森林",
        "研究对象": "GEDI刻画热带森林退化梯度和结构恢复能力",
        "数据源": "GEDI L2A/L2B结构指标、GEDI L4A AGBD；机载LiDAR；MapBiomas/火烧退化类别",
        "GEDI高精度足迹点筛选方法": "使用质量参数过滤低质量回波；具体阈值位于补充说明，正文抽取文本未列明",
        "森林复杂性的计算方法": "使用RH、canopy cover、FHD、PAI、canopy directional gap probability等GEDI指标；PCA提取森林结构状态，PC ratio表征退化连续梯度",
        "空间分辨率": "GEDI足迹尺度；与机载LiDAR足迹级比较",
        "时间范围": "机载LiDAR为2018、2021、2023；火烧/恢复时间涵盖扰动后15-38年等梯度",
        "方法模型": "Lin一致性相关系数；PCA；multinomial logistic regression（MNLR）",
        "评价指标": "CCC；RMSE；rRMSE；Bias；PCA解释率PC1=80%、PC2=15%",
        "主要结论": "GEDI 2A/B指标可区分从原始未火烧到严重退化次生林的结构状态；RH75、RH96、FHD和gap probability对退化梯度重要。",
        "局限性": "相近退化类别存在结构重叠；space-for-time替代有局限；作者建议项目区应用时重新验证PC1/PC2含义。",
        "相关性": "高",
        "相关性判断依据": "直接讨论GEDI结构指标识别森林结构退化和恢复。",
    },
    {
        "PDF文件": "Exploring liana-driven vertical complexity using GEDI simulator and.pdf",
        "标题": "Exploring liana-driven vertical complexity using GEDI simulator and Lorenz-entropy in a neotropical dry forest",
        "作者": "Nooshin Mashhadi; Arturo Sanchez-Azofeifa",
        "年份": "2025",
        "期刊": "International Journal of Applied Earth Observation and Geoinformation",
        "DOI": "10.1016/j.jag.2025.104922",
        "研究区": "哥斯达黎加Santa Rosa National Park Environmental Monitoring SuperSite",
        "研究对象": "藤本植物对热带干旱林垂直结构复杂性的影响",
        "数据源": "机载full-waveform LiDAR；GEDI simulator模拟波形；样地藤本/非藤本及演替阶段信息",
        "GEDI高精度足迹点筛选方法": "使用模拟GEDI波形，不涉及真实GEDI足迹筛选",
        "森林复杂性的计算方法": "Lorenz-entropy（LE）指数；Lorenz曲线/Gini系数量化波形返回能量分布不均衡和垂直异质性",
        "空间分辨率": "GEDI模拟25 m足迹；其他空间分辨率未说明",
        "时间范围": "未说明",
        "方法模型": "GEDI simulator；Mann-Whitney U检验；Benjamini-Hochberg校正；Cliff's delta效应量",
        "评价指标": "P值；BH阈值；Cliff's delta及95% CI",
        "主要结论": "藤本植物对垂直复杂性的影响具有演替阶段依赖性，中/晚期藤本样地LE更高，早期则非藤本样地LE更高。",
        "局限性": "作者明确指出依赖模拟GEDI波形是关键局限；LE对早期林分大冠层空隙敏感。",
        "相关性": "高",
        "相关性判断依据": "虽然是模拟GEDI，但直接探索GEDI波形与垂直复杂性指数。",
    },
    {
        "PDF文件": "Exploring_Forest_Vertical_Structure_With_TomoSense_GEDI_and_SAR_Tomography_Insights.pdf",
        "标题": "Exploring Forest Vertical Structure With TomoSense: GEDI and SAR Tomography Insights",
        "作者": "Yen-Nhi Ngo; Dinh Ho Tong Minh; Nicolas N. Baghdadi; Laurent Ferro-Famil; Yue Huang; Stefano Tebaldini; Ibrahim Fayad",
        "年份": "2025",
        "期刊": "IEEE Transactions on Geoscience and Remote Sensing",
        "DOI": "10.1109/TGRS.2024.3513641",
        "研究区": "德国Eifel National Park的Kermeter区域",
        "研究对象": "GEDI与P-band TomoSAR刻画温带森林垂直结构",
        "数据源": "GEDI；ESA TomoSense airborne P-band TomoSAR；leaf-on/leaf-off数据",
        "GEDI高精度足迹点筛选方法": "抽取文本未说明具体GEDI质量阈值",
        "森林复杂性的计算方法": "以RH指标、DTM/CHM、垂直profile和TomoSAR volume peak分析森林垂直结构；未形成单一复杂性指数",
        "空间分辨率": "未说明",
        "时间范围": "未说明",
        "方法模型": "GEDI与TomoSAR高度/profile对比分析",
        "评价指标": "高度/profile精度指标；摘要未给出具体数值",
        "主要结论": "GEDI与TomoSAR在森林垂直结构测量中互补；该研究建议该温带森林中RH85较RH98更适合树高估计。",
        "局限性": "聚焦单一德国样区；不是复杂性反演专文。",
        "相关性": "中",
        "相关性判断依据": "涉及GEDI垂直结构，但重点是TomoSAR协同与高度/profile。",
    },
    {
        "PDF文件": "Impact_of_GEDI-Derived_Forest_Vertical_Structure_Characteristics_on_the_Accuracy_Gains_in_Regional_Dominant_Tree_Species_Mapping.pdf",
        "标题": "Impact of GEDI-Derived Forest Vertical Structure Characteristics on the Accuracy Gains in Regional Dominant Tree Species Mapping",
        "作者": "Henghui Han; Kaijian Xu; Zhaoying Zhang; Ping Zhao; Hailan Jiang; Anxin Ding",
        "年份": "2024",
        "期刊": "IEEE Geoscience and Remote Sensing Letters",
        "DOI": "10.1109/LGRS.2024.3489214",
        "研究区": "中国东部温带与亚热带森林区",
        "研究对象": "GEDI垂直结构特征对优势树种制图精度的增益",
        "数据源": "GEDI L2A高度和波形强度指标；Sentinel-2；Sentinel-1；PALSAR-2",
        "GEDI高精度足迹点筛选方法": "抽取文本未说明",
        "森林复杂性的计算方法": "将GEDI-derived forest vertical structure characterization（FVSC）作为结构特征；不是专门复杂性指数",
        "空间分辨率": "未说明",
        "时间范围": "覆盖春、夏等物候阶段；具体年份未说明",
        "方法模型": "Random Forest回归反演FVSC；RFE评估特征重要性；Random Forest分层分类树种",
        "评价指标": "FVSC反演R2/RMSE；树种分类精度指标（抽取文本未见具体数值）",
        "主要结论": "GEDI-FVSC补充Sentinel-2光谱特征可提高11种树种制图精度，阔叶树种增益大于针叶树种。",
        "局限性": "某些树种内部结构分化有限，因此GEDI增益受限。",
        "相关性": "中",
        "相关性判断依据": "使用GEDI垂直结构特征，但目标是树种制图而非复杂性反演。",
    },
    {
        "PDF文件": "Influence of GEDI Acquisition and Processing Parameters on Canopy Height Estimates over Tropical Forests.pdf",
        "标题": "Influence of GEDI Acquisition and Processing Parameters on Canopy Height Estimates over Tropical Forests",
        "作者": "Kamel Lahssini; Nicolas Baghdadi; Guerric le Maire; Ibrahim Fayad",
        "年份": "2022",
        "期刊": "Remote Sensing",
        "DOI": "10.3390/rs14246264",
        "研究区": "法属圭亚那两个样区；加蓬两个样区",
        "研究对象": "GEDI采集/处理参数对热带森林冠层高度估计的影响",
        "数据源": "GEDI波形与指标；机载LiDAR CHM参考数据",
        "GEDI高精度足迹点筛选方法": "论文建议使用power beams和高sensitivity beams；具体筛选阈值在抽取文本中未完整列出",
        "森林复杂性的计算方法": "未计算森林复杂性；聚焦冠层高度，讨论GEDI丰富指标可派生结构指标",
        "空间分辨率": "GEDI足迹尺度；ALS CHM分辨率未说明",
        "时间范围": "GEDI自2019年4月起；本研究具体观测期未说明",
        "方法模型": "stepwise multilinear regression；Random Forest；CNN直接使用GEDI波形",
        "评价指标": "RMSE、MAD等；RF相对RH95直接估计RMSE改善约80%",
        "主要结论": "多GEDI指标回归模型优于单一RH95；beam type和beam sensitivity显著影响冠层高度估计。",
        "局限性": "模型迁移性和不同站点/参数配置仍需谨慎；主要服务高度而非复杂性。",
        "相关性": "中",
        "相关性判断依据": "提供GEDI质量控制和高度指标选择经验，可支撑复杂性反演的预处理部分。",
    },
    {
        "PDF文件": "Integration of VIIRS Observations with GEDI-Lidar Measurements to Monitor Forest Structure Dynamics from 2013 to 2020 across the Conterminous United States.pdf",
        "标题": "Integration of VIIRS Observations with GEDI-Lidar Measurements to Monitor Forest Structure Dynamics from 2013 to 2020 across the Conterminous United States",
        "作者": "Khaldoun Rishmawi; Chengquan Huang; Karen Schleeweis; Xiwu Zhan",
        "年份": "2022",
        "期刊": "Remote Sensing",
        "DOI": "10.3390/rs14102320",
        "研究区": "美国本土（CONUS）",
        "研究对象": "2013-2020年森林结构年度动态监测",
        "数据源": "GEDI waveform LiDAR；VIIRS；NEON ALS验证数据；地形/土地覆盖辅助数据",
        "GEDI高精度足迹点筛选方法": "degrade_flag=0；solar elevation<0；sensitivity在tree cover>95%时>0.95，否则>0.9；coverage beams用于tree cover<95%；leaf-off_flag=0；quality_flag=1；urban_proportion=0",
        "森林复杂性的计算方法": "制图CH、PCC、PAI、FHD；FHD作为垂直结构/复杂性指标",
        "空间分辨率": "1 km",
        "时间范围": "2013-2020；GEDI训练样本主要2019和2020",
        "方法模型": "Random Forest regression；VIIRS-GEDI融合年度制图",
        "评价指标": "RMSE；CH RMSE=3.31-4.19 m；PCC RMSE=8%-11%",
        "主要结论": "VIIRS与GEDI融合可生成美国本土年度CH、PCC、PAI、FHD产品，支持多年森林结构动态监测。",
        "局限性": "1 km分辨率较粗；依赖GEDI短期样本训练回推历史年份。",
        "相关性": "高",
        "相关性判断依据": "直接使用GEDI反演/制图FHD等结构复杂性相关指标。",
    },
    {
        "PDF文件": "LiDAR GEDI derived tree canopy height heterogeneity reveals patterns of biodiversity in forest ecosystems.pdf",
        "标题": "LiDAR GEDI derived tree canopy height heterogeneity reveals patterns of biodiversity in forest ecosystems",
        "作者": "Michele Torresani; Duccio Rocchini; Alessandro Alberti; Vítězslav Moudrý; Michael Heym; Elisa Thouverai; Patrick Kacic; Enrico Tomelleri",
        "年份": "2023",
        "期刊": "Ecological Informatics",
        "DOI": "10.1016/j.ecoinf.2023.102082",
        "研究区": "意大利北部阿尔卑斯30个样地；德国Traunstein ForestGEO样地100个样地",
        "研究对象": "GEDI衍生CHM高度异质性与树种多样性",
        "数据源": "GEDI衍生CHM：Potapov30m（GEDI+Landsat）与Lang10m（GEDI+Sentinel-2深度学习）；ALS验证；树种样地数据",
        "GEDI高精度足迹点筛选方法": "引用Potapov30m产品使用power beam、夜间、beam sensitivity 0.9，并排除温带/寒带leaf-off观测；本文自身使用衍生CHM产品",
        "森林复杂性的计算方法": "树冠高度异质性（HH），采用4种异质性指数，包括Rao's Q等；HH作为结构复杂性代理",
        "空间分辨率": "30 m和10 m GEDI衍生CHM；ALS CHM为2.5 m",
        "时间范围": "Potapov30m面向2019；ALS意大利2006、德国2010/2018；样地普查年份见文中",
        "方法模型": "多元回归；交叉验证；比较不同HH指数和CHM产品",
        "评价指标": "R2/相关性；具体数值需查表，抽取文本未完整显示",
        "主要结论": "GEDI衍生CHM可通过高度异质性揭示森林树种多样性格局，但结果受CHM产品、空间分辨率、异质性指数和林分密度影响。",
        "局限性": "方法对数据集选择和空间分辨率敏感；GEDI衍生产品误差会传递到生物多样性解释。",
        "相关性": "高",
        "相关性判断依据": "使用GEDI衍生高度异质性作为结构复杂性代理，适合生态应用章节。",
    },
    {
        "PDF文件": "Mapping global forest canopy height through integration of GEDI and Landsat data.pdf",
        "标题": "Mapping global forest canopy height through integration of GEDI and Landsat data",
        "作者": "Peter Potapov; Xinyuan Li; Andres Hernandez-Serna; Alexandra Tyukavina; Matthew C. Hansen; Anil Kommareddy; Amy Pickens; Svetlana Turubanova; Hao Tang; Carlos Edibaldo Silva; John Armston; Ralph Dubayah; J. Bryan Blair; Michelle Hofton",
        "年份": "2021",
        "期刊": "Remote Sensing of Environment",
        "DOI": "未说明",
        "研究区": "全球森林",
        "研究对象": "全球森林冠层高度制图",
        "数据源": "GEDI足迹级冠层高度；Landsat analysis-ready data",
        "GEDI高精度足迹点筛选方法": "抽取文本未列出详细质量筛选；相关研究常用高质量GEDI足迹，但本文PDF片段未明确",
        "森林复杂性的计算方法": "未计算复杂性；RH/冠层高度可作为复杂性或结构制图基础变量",
        "空间分辨率": "30 m",
        "时间范围": "2019年全球森林冠层高度图",
        "方法模型": "GEDI足迹高度外推到Landsat时序影像；具体模型在抽取文本未完整显示",
        "评价指标": "GEDI验证RMSE=6.6 m、MAE=4.45 m、R2=0.62；ALS验证RMSE=9.07 m、MAE=6.36 m、R2=0.61",
        "主要结论": "GEDI与Landsat融合可生成30 m全球森林冠层高度图，并支持历史和未来森林高度动态监测。",
        "局限性": "高度不是完整复杂性；DOI未在PDF抽取文本中明确显示。",
        "相关性": "中",
        "相关性判断依据": "提供GEDI+Landsat连续制图范式，但目标变量是高度而非复杂性。",
    },
    {
        "PDF文件": "Maps of forest vertical structure for Colombia, a.pdf",
        "标题": "Maps of forest vertical structure for Colombia, a megadiverse country",
        "作者": "J. Camilo Fagua; Patrick Jantz; Patrick Burns; Samuel M. Jantz; John B. Kilbride; Scott J. Goetz",
        "年份": "2025",
        "期刊": "Scientific Data",
        "DOI": "10.1038/s41597-025-06297-7",
        "研究区": "哥伦比亚全国森林，按Andean、Caribbean、Amazon、Chocó、Orinoquía五大自然区建模",
        "研究对象": "哥伦比亚森林垂直结构指标25 m制图",
        "数据源": "超过590万个GEDI足迹；Sentinel-1/2；ALOS-2 PALSAR；多光谱和SAR纹理特征",
        "GEDI高精度足迹点筛选方法": "文中称识别高质量GEDI足迹并用于区域模型；抽取文本未列出具体flag阈值",
        "森林复杂性的计算方法": "制图CH、RH50、total canopy cover、FHD、PAI；FHD和PAI表征垂直结构复杂性/植被量",
        "空间分辨率": "25 m",
        "时间范围": "2020年",
        "方法模型": "区域Random Forest；每区域最多抽样约120万个高质量足迹；82个SAR/多光谱预测变量；Boruta等特征选择",
        "评价指标": "相对误差/模型误差；抽取文本未见完整数值表",
        "主要结论": "5个全国森林结构图保持GEDI足迹中结构指标间关系；Amazon和Andean区域误差较高，total cover相对误差最高。",
        "局限性": "Amazon和Andean区域误差较高；总覆盖度相对误差最大；文章为in press版本。",
        "相关性": "高",
        "相关性判断依据": "直接用GEDI制图FHD/PAI等森林垂直结构复杂性相关变量。",
    },
    {
        "PDF文件": "Modeling Worldwide Tree Biodiversity Using Canopy Structure Metrics from Global Ecosystem Dynamics Investigation Data.pdf",
        "标题": "Modeling Worldwide Tree Biodiversity Using Canopy Structure Metrics from Global Ecosystem Dynamics Investigation Data",
        "作者": "Jin Xu; Kjirsten Coleman; Volker C. Radeloff; Melissa Songer; Qiongyu Huang",
        "年份": "2025",
        "期刊": "Remote Sensing",
        "DOI": "10.3390/rs17081408",
        "研究区": "全球ForestGEO样地，跨气候区",
        "研究对象": "GEDI冠层结构指标预测全球树种丰富度",
        "数据源": "GEDI L2A/L2B；ForestGEO树种丰富度；DHI/NDVI等光谱植被指标；Copernicus森林掩膜",
        "GEDI高精度足迹点筛选方法": "选择2019-2021年5-9月生长季GEDI L2A/L2B；排除quality_flag=0、degrade_flag>1等低质量shot；使用森林覆盖产品掩膜；sensitivity阈值文本抽取存在歧义，具体按原文表述核对",
        "森林复杂性的计算方法": "GEDI指标包括RH98、PAI、cover、FHD、number of layers、PAVD_ratio、PAI_ratio、cover_ratio等均值/标准差；新建4个冠层指数，其中一个代表冠层结构复杂性",
        "空间分辨率": "测试400-6000 m聚合尺度；最优像元/聚合尺度为5600 m",
        "时间范围": "2019、2020、2021年生长季",
        "方法模型": "Random Forest；DHI-only、GEDI-only、GEDI-DHI三类模型；气候区分组建模",
        "评价指标": "R2、RMSE、NRMSE；GEDI-only R2=0.55，GEDI-DHI R2=0.55",
        "主要结论": "GEDI结构指标可较好预测全球树种丰富度；加入DHI未显著提升ForestGEO数据集上的GEDI-only模型表现。",
        "局限性": "气候区差异强，作者建议按气候区建模；需要更多开源地面树种丰富度数据。",
        "相关性": "高",
        "相关性判断依据": "直接利用GEDI结构复杂性相关指标解释生物多样性。",
    },
    {
        "PDF文件": "Remotely sensed 3D ecosystem structure to explain  biodiversity distribution.pdf",
        "标题": "Remotely sensed 3D ecosystem structure to explain biodiversity distribution",
        "作者": "Rezgar Darvand; Omid Esmailzadeh; Habib Zare; Tayebeh Amini; W. Daniel Kissling; Babak Naimi",
        "年份": "2026",
        "期刊": "Ecological Informatics",
        "DOI": "10.1016/j.ecoinf.2026.103770",
        "研究区": "伊朗北部Hyrcanian地区温带落叶林/地中海型群落",
        "研究对象": "GEDI三维生态系统结构指标解释植物分布和群落格局",
        "数据源": "GEDI v2 L2A/L2B（2019-2023）；样地植物群落调查；CHELSA/生物气候变量",
        "GEDI高精度足迹点筛选方法": "排除quality_flag=0、degrade_flag=1；beam sensitivity>0.95；leaf-on season；去除异常值；共准备约1782940个L2A和2719418个L2B footprints",
        "森林复杂性的计算方法": "PAVD、FHD、PAI、RH等；新建3D voxel-based indicators，将多个足迹的水平配置与垂直波形结合，计算体素熵/离散度",
        "空间分辨率": "GEDI约25 m足迹；1×1 km插值/体素尺度；20×20 m样地不直接匹配",
        "时间范围": "GEDI 2019-2023；样地2017-2023",
        "方法模型": "LightGBM特征筛选；RF、SVM、GLM、BRT、Maxent集成SDM；空间/随机交叉验证",
        "评价指标": "AUC；R2；组合模型calibration/evaluation AUC关系R2=0.71，CHELSA R2=0.43，GEDI R2=0.58",
        "主要结论": "GEDI结构指标提升了气候基线之外的物种分布模型，体素3D指标、PAVD和FHD常为重要变量。",
        "局限性": "GEDI地理定位误差使单足迹不宜直接匹配20×20 m样地；陡坡上GEDI波形可能有10-15%高度偏差，但体素聚合可缓解。",
        "相关性": "高",
        "相关性判断依据": "直接使用GEDI三维结构复杂性指标及新体素复杂性指标。",
    },
    {
        "PDF文件": "Stand structure of tropical forests is.pdf",
        "标题": "Stand structure of tropical forests is strongly associated with primary productivity",
        "作者": "Wenmin Zhang; Yanbiao Xi; Martin Brandt; Chunying Ren; Jialing Bai; Qin Ma; Rasmus Fensholt",
        "年份": "2024",
        "期刊": "Communications Earth & Environment",
        "DOI": "10.1038/s43247-024-01984-6",
        "研究区": "全球热带森林",
        "研究对象": "热带森林林分结构指标SSI与初级生产力",
        "数据源": "GEDI结构指标；Sentinel-1/2；生产力相关数据",
        "GEDI高精度足迹点筛选方法": "抽取文本未说明具体足迹筛选阈值",
        "森林复杂性的计算方法": "基于4个GEDI结构指标用PCA推导Stand Structural Indicator（SSI），同时捕捉高度、密度和冠层高度多样性",
        "空间分辨率": "GEDI足迹尺度并外推为连续分布；具体栅格分辨率未在抽取片段中说明",
        "时间范围": "未说明",
        "方法模型": "PCA；Sentinel-1/2上尺度制图；生产力相关分析",
        "评价指标": "相关性/模型统计；具体数值需查原文表格",
        "主要结论": "高SSI主要分布于湿润、天然、低火扰动热带森林；SSI与初级生产力正相关，天然林中生产力对SSI更敏感。",
        "局限性": "具体GEDI筛选和分辨率信息需回查方法表；SSI是综合指标，生态含义需按输入变量解释。",
        "相关性": "高",
        "相关性判断依据": "直接以GEDI构建综合结构指标并外推，接近结构复杂性反演。",
    },
    {
        "PDF文件": "Stratifying forest overstory and understory using the Global Ecosystem.pdf",
        "标题": "Stratifying forest overstory and understory using the Global Ecosystem Dynamic Investigation laser scanning data",
        "作者": "Zengxin Yun; Guang Zheng; L. Monika Moskal; Jiarui Li; Peng Gong",
        "年份": "2023",
        "期刊": "International Journal of Applied Earth Observation and Geoinformation",
        "DOI": "10.1016/j.jag.2023.103538",
        "研究区": "主要验证于Panther Creek等样区；抽取文本显示实验验证集中在Panther Creek",
        "研究对象": "使用GEDI L1B分离森林上层和下层波形",
        "数据源": "GEDI L1B；ALS验证数据",
        "GEDI高精度足迹点筛选方法": "抽取文本未说明完整筛选阈值；方法考虑多峰波形、冠层覆盖和坡度",
        "森林复杂性的计算方法": "未构建复杂性指数；通过上层/下层波形分层、冠层覆盖、地形坡度等识别多层结构",
        "空间分辨率": "GEDI足迹尺度",
        "时间范围": "未说明",
        "方法模型": "多峰波形分解；overstory/understory提取算法；ALS对比",
        "评价指标": "overstory height bias/RMSE；understory height bias/RMSE；ALS-GEDI canopy cover R2",
        "主要结论": "方法改善GEDI上层高度偏差和RMSE；下层高度bias=2.85 m、RMSE=3.83 m；地形、冠层覆盖、高度和定位误差影响分层。",
        "局限性": "作者指出实验和验证主要在Panther Creek，未来需区域/全球应用；陡坡森林需进一步算法改进。",
        "相关性": "中",
        "相关性判断依据": "服务GEDI垂直分层和复杂结构解析，但不直接反演复杂性指标。",
    },
    {
        "PDF文件": "The use of GEDI canopy structure for explaining variation in tree.pdf",
        "标题": "The use of GEDI canopy structure for explaining variation in tree species richness in natural forests",
        "作者": "Suzanne M. Marselis; Petr Keil; Jonathan M. Chase; Ralph Dubayah",
        "年份": "2022",
        "期刊": "Environmental Research Letters",
        "DOI": "10.1088/1748-9326/ac583f",
        "研究区": "全球自然/半自然森林样地",
        "研究对象": "GEDI冠层结构解释树种丰富度变化",
        "数据源": "GEDI L2A/L2B canopy metrics；全球森林样地树种丰富度；气候、地形、生物地理变量",
        "GEDI高精度足迹点筛选方法": "使用leaf-on季节；仅使用L2B successful processing flag=1的可靠波形；按1、4、16 km2缓冲区汇总高质量足迹",
        "森林复杂性的计算方法": "使用RH98、total PAI、PAVD垂直分层、FHD、canopy cover等9个结构指标及其均值/标准差/最大/最小；未形成单一复杂性指数",
        "空间分辨率": "1 km2、4 km2、16 km2缓冲区聚合",
        "时间范围": "GEDI任务期；具体观测年份未说明",
        "方法模型": "Random Forest；PCA辅助筛选指标；variation partitioning",
        "评价指标": "R2/解释变异；GEDI结构最多解释66%树种丰富度变异",
        "主要结论": "GEDI冠层结构与自然/半自然未扰动森林树种丰富度显著相关，但该关系与环境/生物地理变量解释部分重叠。",
        "局限性": "数据集不含严重扰动林、人工林和单一树种种植林；GEDI采样密度不足以研究20×20 m样地尺度。",
        "相关性": "高",
        "相关性判断依据": "直接使用GEDI结构指标解释生物多样性，是生态应用重要文献。",
    },
    {
        "PDF文件": "Upscaling eddy covariance measurements of carbon and water1fluxes to the continental scale by incorporating GEDI-derived2canopy structural metrics3.pdf",
        "标题": "Upscaling eddy covariance measurements of carbon and water fluxes to the continental scale by incorporating GEDI-derived canopy structural metrics",
        "作者": "Jingyi Bu; Jingfeng Xiao",
        "年份": "未说明",
        "期刊": "未说明",
        "DOI": "未说明",
        "研究区": "美国本土（CONUS）；AmeriFlux和NEON站点",
        "研究对象": "GEDI冠层结构指标改进GPP和ET通量上尺度",
        "数据源": "GEDI02_B FHD、RH、CC、PAI、PGAP；AmeriFlux/NEON通量塔；MODIS NIRv；Daymet气象；土壤水分等",
        "GEDI高精度足迹点筛选方法": "sensitivity 0.90-1；algorithmrun_flag=1；stale_return_flag=0；RH限制0-12000；l2b_quality_flag=1；elev_lowestmode -200到9000；surface_flag=1；degrade_flag取指定集合；站点1 km半径内至少60个有效GEDI足迹",
        "森林复杂性的计算方法": "FHD表征冠层层数和植物物质垂直分布均匀度；同时使用RH/CC/PAI/PGAP；以FHD分析复杂性对GPP/ET与抗旱性的影响",
        "空间分辨率": "1 km；日尺度GPP/ET数据集",
        "时间范围": "GEDI 2019年4月-2023年3月；输出2019-2023",
        "方法模型": "XGBoost；70/30训练验证；5-fold cross validation",
        "评价指标": "GPP R2从0.79升至0.91、RMSE从1.77降至1.14 gC m-2 d-1；ET R2从0.79升至0.85、RMSE从0.82降至0.68 mm d-1",
        "主要结论": "纳入GEDI RH和FHD显著提高GPP/ET上尺度模型；FHD越高，GPP/ET总体增加但边际效应减弱，且复杂冠层降低季节变异并增强干旱韧性。",
        "局限性": "PDF未显示期刊/DOI；结构-功能关系还需结合气候、土壤和生物因素综合解释。",
        "相关性": "高",
        "相关性判断依据": "直接将GEDI FHD作为复杂性指标用于碳水通量模型。",
    },
    {
        "PDF文件": "Vertical complexity matters How canopy height and architecture mitigate.pdf",
        "标题": "Vertical complexity matters: How canopy height and architecture mitigate urban heat in high-density environments",
        "作者": "Hongling Yi; Fei Feng; Xianwen Li; Chengyang Xu; Yongxian Su; Peng Yao; Raffaele Lafortezza",
        "年份": "2026",
        "期刊": "Urban Forestry & Urban Greening",
        "DOI": "10.1016/j.ufug.2026.129506",
        "研究区": "北京高密度城市环境",
        "研究对象": "城市森林垂直结构对地表温度/热岛缓解的作用",
        "数据源": "GEDI LiDAR；Landsat LST；城市冠层/覆盖相关数据",
        "GEDI高精度足迹点筛选方法": "抽取文本未说明具体筛选方法",
        "森林复杂性的计算方法": "使用RH95等垂直结构指标，以及覆盖度、PAI等水平/体量指标；未见单一复杂性指数",
        "空间分辨率": "未说明",
        "时间范围": "未说明",
        "方法模型": "Structural Equation Modeling（SEM）",
        "评价指标": "SEM路径效应；摘要称RH95约降低LST 1.13°C",
        "主要结论": "垂直结构（RH95）是城市热缓解的重要因素，相比二维覆盖更能解释降温贡献。",
        "局限性": "城市森林场景，与自然森林结构复杂性反演主题有生态场景差异。",
        "相关性": "中",
        "相关性判断依据": "使用GEDI结构指标，但对象是城市热环境而非森林结构复杂性制图。",
    },
    {
        "PDF文件": "利用冠层三维结构特征改进瑞士阿尔高州星载光子计数激光雷达森林蓄积量估测_NormalPdf.pdf",
        "标题": "利用冠层三维结构特征改进瑞士阿尔高州星载光子计数激光雷达森林蓄积量估测",
        "作者": "孔丹; 庞勇; 汪祖媛; 李增元",
        "年份": "2025",
        "期刊": "遥感学报 / National Remote Sensing Bulletin",
        "DOI": "10.11834/jrs.20254317",
        "研究区": "瑞士阿尔高州针阔混交森林",
        "研究对象": "ICESat-2 ATLAS森林蓄积量估测中的三维冠层结构特征",
        "数据源": "ICESat-2 ATLAS；森林蓄积量参考数据；结构特征",
        "GEDI高精度足迹点筛选方法": "不适用（使用ICESat-2而非GEDI）",
        "森林复杂性的计算方法": "冠高及冠高异质性、垂直结构特征、叶面积加权冠层体积、枝叶剖面等三维结构特征",
        "空间分辨率": "100 m估测单元",
        "时间范围": "未说明",
        "方法模型": "质量控制；特征分组预筛选；有规则约束的全子集筛选；十折交叉验证",
        "评价指标": "R2=0.78，RMSE=92.48 m3/hm2，rRMSE=0.24；基线R2=0.66、rRMSE=0.28",
        "主要结论": "引入冠层三维结构特征可提升ICESat-2复杂林分蓄积量估测精度。",
        "局限性": "不是GEDI研究；结果针对瑞士阿尔高州针阔混交林。",
        "相关性": "低",
        "相关性判断依据": "可借鉴结构特征构造和光子计数激光雷达质量控制，但不属于GEDI应用。",
    },
    {
        "PDF文件": "星载激光遥感林业应用发展研究.pdf",
        "标题": "星载激光遥感林业应用发展研究",
        "作者": "岳春宇; 郑永超; 邢艳秋; 庞勇; 李世明; 蔡龙涛; 何红艳",
        "年份": "2020",
        "期刊": "红外与激光工程 / Infrared and Laser Engineering",
        "DOI": "10.3788/IRLA20200235",
        "研究区": "综述文献，未限定单一研究区",
        "研究对象": "星载LiDAR林业遥感技术与应用发展",
        "数据源": "ICESat/GLAS、ICESat-2/ATLAS、GEDI等星载激光雷达任务相关研究",
        "GEDI高精度足迹点筛选方法": "综述文献，未提供实证筛选方法",
        "森林复杂性的计算方法": "综述星载激光雷达在树高、森林类型、生物量、垂直结构等方面应用；未计算复杂性",
        "空间分辨率": "不同任务不同；文中提到光斑变小、波束增多、频率提高的发展趋势",
        "时间范围": "截至2020年前后星载LiDAR任务",
        "方法模型": "综述",
        "评价指标": "综述若干前人研究精度，如GLAS树高/类型分类等；未形成统一评价",
        "主要结论": "星载LiDAR逐步由大光斑、低频向小光斑、多波束和高频发展，可获取森林垂直结构并支撑生物量/碳循环估计。",
        "局限性": "宏观观测并非采样越密精度越高；地形、信噪比、覆盖度和多源协同仍是挑战。",
        "相关性": "中",
        "相关性判断依据": "GEDI技术背景和星载LiDAR发展综述，可用于绪论。",
    },
    {
        "PDF文件": "森林冠层结构复杂性研究进展及展望_NormalPdf.pdf",
        "标题": "森林冠层结构复杂性研究进展及展望",
        "作者": "胡天宇; 刘小强; 吴晓永; 牛春跃; 苏艳军",
        "年份": "2025",
        "期刊": "遥感学报 / National Remote Sensing Bulletin",
        "DOI": "10.11834/jrs.20244007",
        "研究区": "综述文献，未限定单一研究区",
        "研究对象": "森林冠层结构复杂性概念、量化方法与生态应用",
        "数据源": "地面/近地面/机载/星载LiDAR与相关生态研究文献",
        "GEDI高精度足迹点筛选方法": "不适用（综述文献，未做GEDI足迹筛选）",
        "森林复杂性的计算方法": "总结树高/胸径变异、Gini、冠层熵、分形维数、SSCI、水平/垂直/整合分布指标等；强调LiDAR三维结构信息",
        "空间分辨率": "视平台而定；综述指出尺度和平台通用性是关键问题",
        "时间范围": "截至2025年前相关研究",
        "方法模型": "综述",
        "评价指标": "未说明",
        "主要结论": "冠层结构复杂性已广泛用于光、水、微气候、生产力和稳定性研究；LiDAR促进其准确三维量化。",
        "局限性": "作者指出指标含义、跨平台通用性、尺度依赖和长时间序列监测仍是未来重点难题。",
        "相关性": "高",
        "相关性判断依据": "虽然非GEDI实证，但直接提供森林结构复杂性的理论和指标框架，是章节写作关键综述。",
    },
]


columns = [
    "PDF文件",
    "标题",
    "作者",
    "年份",
    "期刊",
    "DOI",
    "研究区",
    "研究对象",
    "数据源",
    "GEDI高精度足迹点筛选方法",
    "森林复杂性的计算方法",
    "空间分辨率",
    "时间范围",
    "方法模型",
    "评价指标",
    "主要结论",
    "局限性",
    "相关性",
    "相关性判断依据",
]


def write_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "文献编码表"
    ws.append(columns)

    for row in rows:
        ws.append([row.get(col, "未说明") or "未说明" for col in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 42, "B": 48, "C": 42, "D": 10, "E": 26, "F": 28,
        "G": 34, "H": 34, "I": 44, "J": 52, "K": 54, "L": 22,
        "M": 24, "N": 36, "O": 32, "P": 56, "Q": 48, "R": 12, "S": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    note = wb.create_sheet("字段说明")
    notes = [
        ("字段", "说明"),
        ("未说明", "表示未能从PDF正文、首页题录或抽取文本中可靠确定。"),
        ("不适用", "表示文献明确未使用GEDI，或为综述/非GEDI传感器研究。"),
        ("相关性：高", "直接使用GEDI进行森林结构复杂性、垂直结构、FHD/PAVD/PAI等反演、制图或生态应用。"),
        ("相关性：中", "涉及LiDAR结构复杂性理论、GEDI预处理/高度/分层，或非森林复杂性主应用。"),
        ("相关性：低", "不使用GEDI且仅能间接借鉴。"),
    ]
    for item in notes:
        note.append(item)
    for cell in note[1]:
        cell.fill = header_fill
        cell.font = header_font
    note.column_dimensions["A"].width = 22
    note.column_dimensions["B"].width = 90
    for row in note.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap

    wb.save(OUT_XLSX)


def write_summary():
    high = [r for r in rows if r["相关性"] == "高"]
    medium = [r for r in rows if r["相关性"] == "中"]
    low = [r for r in rows if r["相关性"] == "低"]
    md = f"""# GEDI数据在森林结构复杂性反演中的应用：文献整理总结

## 整理范围

本次共整理PDF文献 {len(rows)} 篇，其中与“GEDI数据在森林结构复杂性反演中的应用”高度相关 {len(high)} 篇，中等相关 {len(medium)} 篇，低相关 {len(low)} 篇。字段中凡未能从PDF可靠确定的信息均标注为“未说明”；未使用GEDI的文献在GEDI足迹筛选字段中标注为“不适用”。

## 研究热点

1. **从GEDI稀疏足迹到连续结构复杂性制图。** 典型路线是以GEDI足迹上的RH、cover、PAI、PAVD、FHD或综合复杂性指数为响应变量，再融合Sentinel、Landsat、VIIRS、SAR、地形和气候数据进行上尺度制图。欧洲森林结构多样性数据集使用Random Forest生成1/5/10 km结构多样性产品（Girardello et al., 2026）；哥伦比亚研究用590多万个GEDI足迹和SAR/多光谱影像生成25 m国家尺度结构图（Fagua et al., 2025）；美国西部野生动物生境研究生成30 m GEDI-fusion结构指标图（Vogeler et al., 2023）。

2. **构建可解释的结构复杂性综合指标。** 全球森林WSCI研究将GEDI RH metrics与ALS三维复杂性CE_XYZ连接，提出GEDI Waveform Structural Complexity Index（de Conto et al., 2024）。全球生产力稳定性研究采用canopy entropy并用GEDI与机载LiDAR估计全球CSC（Liu et al., 2024）。热带森林研究用PCA把4个GEDI结构指标合成为SSI（Zhang et al., 2024）。亚马逊退化研究用PCA和PC ratio表征森林结构状态连续梯度（Doyle et al., 2025）。

3. **FHD、PAVD、PAI成为复杂性/垂直结构应用的高频变量。** VIIRS-GEDI研究制图CH、PCC、PAI和FHD（Rishmawi et al., 2022）；美国通量上尺度研究表明加入GEDI RH和FHD可显著提升GPP/ET模型表现（Bu & Xiao, 未说明年份）；伊朗生物多样性研究中PAVD、FHD和3D体素指标常为重要预测变量（Darvand et al., 2026）。

4. **结构复杂性与生态功能、生物多样性和生境质量耦合。** GEDI结构指标被用于解释树种丰富度（Marselis et al., 2022；Xu et al., 2025）、森林生物多样性格局（Torresani et al., 2023；Darvand et al., 2026）、野生动物生境（Vogeler et al., 2023）、生产力和稳定性（Liu et al., 2024；Zhang et al., 2024；Bu & Xiao, 未说明年份）。

5. **GEDI质量控制与参数敏感性是方法基础。** 多篇文献强调degrade flag、quality flag、sensitivity、leaf-on/leaf-off、solar elevation、beam power等筛选条件。Vogeler et al. (2023) 使用solar elevation<0、degrade_flag=0、quality_flag=1、sensitivity>=0.95和full power beams；Rishmawi et al. (2022) 使用degrade_flag=0、solar elevation<0、leaf-off_flag=0、quality_flag=1等；Lahssini et al. (2022) 指出power/high-sensitivity beams对热带森林冠层高度估计非常关键。

## 常用数据源

- **GEDI产品：** L1B波形、L2A RH高度指标、L2B canopy cover/PAI/PAVD/FHD、L4A AGBD。
- **光学遥感：** Landsat、Sentinel-2、VIIRS、MODIS NIRv/NDVI/GPP、GOSIF。
- **雷达/SAR：** Sentinel-1、ALOS-2 PALSAR、TomoSAR/TomoSense。
- **验证与辅助数据：** ALS点云、UAV LiDAR、ForestGEO样地、AmeriFlux/NEON通量塔、NEON ALS、地形/气候/土壤/扰动产品。

## 主要方法

- **机器学习上尺度：** Random Forest最常见，应用于欧洲结构多样性制图、美国西部GEDI-fusion、哥伦比亚25 m结构图、VIIRS-GEDI年度结构制图、全球树种丰富度建模等。
- **深度学习/梯度提升：** XGBoost用于GPP/ET通量上尺度（Bu & Xiao）；CNN直接处理GEDI波形用于冠层高度估计（Lahssini et al., 2022）；LightGBM用于GEDI结构变量筛选（Darvand et al., 2026）。
- **综合指数构建：** WSCI、canopy entropy、SSI、Lorenz-entropy、PCA-derived forest state/PCR、height heterogeneity（Rao's Q等）。
- **统计与生态模型：** PCA、MNLR、SEM、SDM ensemble、variation partitioning、Lin's CCC、Mann-Whitney U和BH校正。

## 研究空白

1. **复杂性定义仍不统一。** WSCI、canopy entropy、FHD、PAVD、SSI、HH、Lorenz-entropy都可表征复杂性，但它们对应的是不同结构维度，跨研究可比性不足（de Conto et al., 2024；胡天宇等, 2025）。

2. **足迹尺度到连续栅格存在尺度不匹配。** GEDI约25 m足迹与30 m、25 m、1 km、5.6 km等栅格或样地尺度之间存在聚合和定位误差问题。Darvand et al. (2026) 明确指出GEDI定位误差使单足迹不宜直接匹配20×20 m样地。

3. **高质量足迹筛选标准尚未完全一致。** 不同研究对sensitivity、beam power、leaf-on/off、solar elevation、degrade flag的阈值不同，导致结果可比性和复现性需要特别说明。

4. **热带、温带、城市、人工林和退化林机制差异明显。** de Conto et al. (2024) 显示热带和温带森林中不同垂直层对复杂性的贡献不同；Doyle et al. (2025) 显示相近退化类别存在结构重叠。

5. **时间序列复杂性监测仍不足。** GEDI观测期有限，长时间序列通常需要VIIRS/Landsat/MODIS等历史数据融合。Rishmawi et al. (2022) 和Bu & Xiao提供了回推/上尺度思路，但复杂性本身的长期变化仍有空间。

## 可写选题方向

1. **基于GEDI L2A/L2B与Sentinel-1/2的区域森林结构复杂性连续制图。** 可借鉴Vogeler et al. (2023)、Fagua et al. (2025)和Girardello et al. (2026)，重点比较FHD/PAVD/PAI与综合指数的表现。

2. **GEDI足迹质量筛选对结构复杂性反演精度的影响。** 可围绕sensitivity、beam type、solar elevation、degrade_flag、leaf-on/off做敏感性分析，连接Lahssini et al. (2022)、Rishmawi et al. (2022)和Vogeler et al. (2023)。

3. **面向森林退化/恢复监测的GEDI结构状态指数构建。** 可参考Doyle et al. (2025)的PCA/PC ratio方法，并结合WSCI或canopy entropy构建更稳健的退化梯度指标。

4. **GEDI结构复杂性指标对生物多样性或生态功能的解释力比较。** 比较RH、FHD、PAVD、PAI、HH、WSCI、canopy entropy对树种丰富度、GPP/ET或生境模型的贡献，参考Marselis et al. (2022)、Xu et al. (2025)、Darvand et al. (2026)和Bu & Xiao。

5. **多尺度结构复杂性反演：足迹、样地、景观与区域尺度的一致性。** 解决GEDI足迹与样地/栅格尺度不匹配问题，特别适合章节中作为“方法挑战与展望”展开。

## 主题图

```mermaid
flowchart LR
    A["GEDI L1B/L2A/L2B/L4A"] --> B["足迹质量筛选"]
    B --> C["结构指标: RH, Cover, PAI, PAVD, FHD"]
    C --> D["复杂性指数: WSCI, Canopy Entropy, SSI, LE, HH"]
    D --> E["多源融合: Landsat, Sentinel, VIIRS, SAR, ALS"]
    E --> F["连续制图: 25 m, 30 m, 1 km, 多尺度"]
    F --> G["生态应用: 生物多样性, 生产力, 退化恢复, 生境质量"]
```
"""
    OUT_MD.write_text(md, encoding="utf-8")


if __name__ == "__main__":
    write_xlsx()
    write_summary()
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MD}")
